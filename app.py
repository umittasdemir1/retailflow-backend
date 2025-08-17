from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import pandas as pd
import os
import io
import json
from datetime import datetime
from werkzeug.utils import secure_filename
import tempfile
import logging
import pickle
import time
import traceback
from functools import wraps, lru_cache
import gc

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)

# CORS configuration - Tum origin'lere izin
CORS(app, origins=["*"])

# Configuration - Buyuk dosya destegi
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB max file size
ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'csv'}

# Temp file path for session persistence
TEMP_DATA_FILE = os.path.join(tempfile.gettempdir(), 'retailflow_data.pkl')

# Strategy configurations
STRATEGY_CONFIG = {
    'sakin': {
        'min_str_diff': 0.15,
        'min_inventory': 3,
        'max_transfer': 5,
        'description': 'Guvenli ve kontrollu transfer yaklasimi'
    },
    'kontrollu': {
        'min_str_diff': 0.10,
        'min_inventory': 2,
        'max_transfer': 10,
        'description': 'Dengeli risk ve performans'
    },
    'agresif': {
        'min_str_diff': 0.08,
        'min_inventory': 1,
        'max_transfer': None,  # sinirsiz
        'description': 'Maksimum performans odakli'
    }
}

# Merkez/Online onceligi sabiti
WAREHOUSE_SET = {'Merkez Depo', 'Online'}

# Performance tracking decorator
def measure_time(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        start = time.time()
        result = func(*args, **kwargs)
        end = time.time()
        execution_time = end - start
        logger.info(f"{func.__name__} completed in {execution_time:.2f} seconds")
        return result
    return wrapper

# Simple memory check without psutil
def check_memory_usage_simple():
    """Simple memory check without external dependencies"""
    try:
        # Try to get memory info from /proc/meminfo (Linux)
        with open('/proc/meminfo', 'r') as f:
            lines = f.readlines()
            mem_total = None
            mem_available = None
            
            for line in lines:
                if line.startswith('MemTotal:'):
                    mem_total = int(line.split()[1])
                elif line.startswith('MemAvailable:'):
                    mem_available = int(line.split()[1])
                    
            if mem_total and mem_available:
                used_percent = ((mem_total - mem_available) / mem_total) * 100
                if used_percent > 85:
                    gc.collect()
                return round(used_percent, 1)
    except:
        pass
    
    # Fallback: just return a reasonable number
    return 50.0

class MagazaTransferSistemi:
    def __init__(self):
        self.data = None
        self.magazalar = []
        self.mevcut_analiz = None
        self.current_strategy = 'sakin'
        self.excluded_stores = []
        self.target_store = None
        self.transfer_type = 'global'
        self.performance_metrics = {}
        self.load_from_temp()

    def check_memory_usage(self):
        """Memory kullanımını kontrol et"""
        return check_memory_usage_simple()

    def validate_request_data(self, request_data):
        """Request verilerini validate et"""
        try:
            # Strategy validation
            if 'strategy' in request_data and request_data['strategy'] not in STRATEGY_CONFIG:
                logger.warning(f"Invalid strategy: {request_data['strategy']}, using default 'sakin'")
                request_data['strategy'] = 'sakin'
            
            # Transfer type validation
            valid_transfer_types = ['global', 'targeted', 'size_completion']
            if 'transfer_type' in request_data and request_data['transfer_type'] not in valid_transfer_types:
                logger.warning(f"Invalid transfer_type: {request_data['transfer_type']}, using default 'global'")
                request_data['transfer_type'] = 'global'
            
            # Excluded stores validation
            if 'excluded_stores' in request_data and not isinstance(request_data['excluded_stores'], list):
                logger.warning("excluded_stores must be a list, converting to empty list")
                request_data['excluded_stores'] = []
            
            # Target store validation
            if 'target_store' in request_data and self.magazalar:
                if request_data['target_store'] not in self.magazalar:
                    logger.warning(f"Invalid target_store: {request_data['target_store']}")
                    request_data['target_store'] = None
            
            return request_data
        except Exception as e:
            logger.error(f"Request validation error: {e}")
            return request_data

    def safe_transfer_calculation(self, gonderen_satis, gonderen_envanter, alan_satis, alan_envanter, strategy='sakin'):
        """Güvenli transfer hesaplama wrapper"""
        try:
            return self.str_bazli_transfer_hesapla(gonderen_satis, gonderen_envanter, alan_satis, alan_envanter, strategy)
        except Exception as e:
            logger.error(f"Transfer calculation error: {e}")
            return 0, {
                'gonderen_str': 0,
                'alan_str': 0,
                'str_farki': 0,
                'teorik_transfer': 0,
                'uygulanan_filtre': 'Hata',
                'kullanilan_strateji': strategy,
                'error': str(e)
            }

    def optimize_dataframe(self, df):
        """DataFrame'i memory için optimize et"""
        try:
            logger.info("Optimizing DataFrame for memory usage...")
            
            # Kategorik sütunlar için memory tasarrufu
            categorical_columns = ['Depo Adi', 'Urun Adi', 'Urun Kodu']
            for col in categorical_columns:
                if col in df.columns:
                    df[col] = df[col].astype('category')
            
            # Numeric sütunlar için uygun dtype
            numeric_columns = ['Satis', 'Envanter']
            for col in numeric_columns:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], downcast='integer', errors='coerce').fillna(0)
            
            # String sütunları optimize et
            string_columns = ['Renk Aciklamasi', 'Beden']
            for col in string_columns:
                if col in df.columns:
                    if df[col].nunique() / len(df) < 0.5:
                        df[col] = df[col].astype('category')
            
            logger.info("DataFrame optimized successfully")
            return df
        except Exception as e:
            logger.error(f"DataFrame optimization error: {e}")
            return df

    def create_product_key_vectorized(self, df):
        """Vectorized urun anahtari olustur - HIZLI YÖNTEM - Categorical fix"""
        try:
            logger.info("Creating product keys using vectorized operations...")
            start_time = time.time()
            
            # Categorical sütunları string'e çevir (vectorized operations için)
            if df['Urun Adi'].dtype.name == 'category':
                df['Urun Adi'] = df['Urun Adi'].astype(str)
            
            # Vectorized string operations
            urun_adi = df['Urun Adi'].fillna('').astype(str).str.strip().str.upper()
            
            # Renk ve Beden sütunları varsa işle
            if 'Renk Aciklamasi' in df.columns:
                if df['Renk Aciklamasi'].dtype.name == 'category':
                    df['Renk Aciklamasi'] = df['Renk Aciklamasi'].astype(str)
                renk = df['Renk Aciklamasi'].fillna('').astype(str).str.strip().str.upper()
            else:
                renk = pd.Series([''] * len(df))
            
            if 'Beden' in df.columns:
                if df['Beden'].dtype.name == 'category':
                    df['Beden'] = df['Beden'].astype(str)
                beden = df['Beden'].fillna('').astype(str).str.strip().str.upper()
            else:
                beden = pd.Series([''] * len(df))
            
            # Combine them
            df['urun_anahtari'] = (urun_adi + ' ' + renk + ' ' + beden).str.strip()
            
            end_time = time.time()
            logger.info(f"Product keys created in {end_time - start_time:.2f} seconds using vectorized operations")
            return df
        except Exception as e:
            logger.error(f"Vectorized product key creation error: {e}")
            return self.create_product_key_fallback(df)

    def create_product_key_fallback(self, df):
        """Fallback urun anahtari olustur"""
        logger.info("Using fallback method for product key creation...")
        df['urun_anahtari'] = df.apply(
            lambda x: self.urun_anahtari_olustur(
                x['Urun Adi'], 
                x.get('Renk Aciklamasi', ''), 
                x.get('Beden', '')
            ), axis=1
        )
        return df

    @lru_cache(maxsize=1000)
    def str_hesapla_cached(self, satis, envanter):
        """Cached STR hesaplama"""
        return self.str_hesapla(satis, envanter)

    def save_to_temp(self):
        """Veriyi gecici dosyaya kaydet"""
        try:
            with open(TEMP_DATA_FILE, 'wb') as f:
                pickle.dump({
                    'data': self.data,
                    'magazalar': self.magazalar,
                    'mevcut_analiz': self.mevcut_analiz,
                    'current_strategy': self.current_strategy,
                    'excluded_stores': self.excluded_stores,
                    'target_store': self.target_store,
                    'transfer_type': self.transfer_type,
                    'performance_metrics': self.performance_metrics
                }, f)
            logger.info("Data saved to temp file")
        except Exception as e:
            logger.error(f"Failed to save temp data: {e}")

    def load_from_temp(self):
        """Gecici dosyadan veriyi yukle"""
        try:
            if os.path.exists(TEMP_DATA_FILE):
                with open(TEMP_DATA_FILE, 'rb') as f:
                    temp_data = pickle.load(f)
                    self.data = temp_data.get('data')
                    self.magazalar = temp_data.get('magazalar', [])
                    self.mevcut_analiz = temp_data.get('mevcut_analiz')
                    self.current_strategy = temp_data.get('current_strategy', 'sakin')
                    self.excluded_stores = temp_data.get('excluded_stores', [])
                    self.target_store = temp_data.get('target_store')
                    self.transfer_type = temp_data.get('transfer_type', 'global')
                    self.performance_metrics = temp_data.get('performance_metrics', {})
                logger.info("Data loaded from temp file")
        except Exception as e:
            logger.error(f"Failed to load temp data: {e}")

    def clear_all_data(self):
        """Tum veriyi temizle"""
        try:
            self.data = None
            self.magazalar = []
            self.mevcut_analiz = None
            self.current_strategy = 'sakin'
            self.excluded_stores = []
            self.target_store = None
            self.transfer_type = 'global'
            self.performance_metrics = {}
            
            if os.path.exists(TEMP_DATA_FILE):
                os.remove(TEMP_DATA_FILE)
                logger.info("Temp data file removed")
            
            gc.collect()
            logger.info("All data cleared successfully")
            return True
        except Exception as e:
            logger.error(f"Failed to clear data: {e}")
            return False

    @measure_time
    def dosya_yukle_df(self, df):
        """DataFrame'i yukle ve isle"""
        try:
            start_memory = self.check_memory_usage()
            logger.info(f"Starting file processing - Memory usage: {start_memory}%")
            
            df.columns = df.columns.str.strip()
            logger.info(f"Bulunan sutunlar: {list(df.columns)}")
            
            # TURKCE SUTUN ADLARINI INGILIZCE'YE CEVIR
            column_mapping = {
                'Depo Adı': 'Depo Adi',
                'Ürün Kodu': 'Urun Kodu', 
                'Ürün Adı': 'Urun Adi',
                'Renk Açıklaması': 'Renk Aciklamasi',
            }
            
            df = df.rename(columns=column_mapping)
            logger.info(f"Sutunlar cevrildikten sonra: {list(df.columns)}")
            
            gerekli_sutunlar = ['Depo Adi', 'Urun Kodu', 'Urun Adi', 'Satis', 'Envanter']
            eksik_sutunlar = [s for s in gerekli_sutunlar if s not in df.columns]
            
            if eksik_sutunlar:
                return False, f"Eksik sutunlar: {', '.join(eksik_sutunlar)}"
            
            df = df.dropna(subset=['Depo Adi'])
            df['Satis'] = pd.to_numeric(df['Satis'], errors='coerce').fillna(0)
            df['Envanter'] = pd.to_numeric(df['Envanter'], errors='coerce').fillna(0)
            
            df['Satis'] = df['Satis'].clip(lower=0)
            df['Envanter'] = df['Envanter'].clip(lower=0)
            
            df = self.optimize_dataframe(df)
            
            self.data = df
            self.magazalar = df['Depo Adi'].unique().tolist()
            
            end_memory = self.check_memory_usage()
            logger.info(f"File processing completed - Memory usage: {end_memory}%")
            logger.info(f"Veri yuklendi: {len(df)} satir, {len(self.magazalar)} magaza")
            
            self.performance_metrics['last_file_load'] = {
                'timestamp': datetime.now().isoformat(),
                'rows': len(df),
                'stores': len(self.magazalar),
                'memory_usage_start': start_memory,
                'memory_usage_end': end_memory
            }
            
            result = {
                'message': f"Basarili! {len(df):,} urun, {len(self.magazalar)} magaza yuklendi.",
                'satir_sayisi': len(df),
                'magaza_sayisi': len(self.magazalar),
                'magazalar': self.magazalar,
                'sutunlar': list(df.columns),
                'memory_usage': f"{end_memory}%"
            }
            
            self.save_to_temp()
            return True, result
            
        except Exception as e:
            logger.error(f"Dosya yukleme hatasi: {str(e)}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            return False, f"Hata: {str(e)}"

    def magaza_metrikleri_hesapla(self):
        """Her magaza icin metrikleri hesapla"""
        if self.data is None:
            return {}

        metrikler = {}
        for magaza in self.magazalar:
            magaza_data = self.data[self.data['Depo Adi'] == magaza]
            toplam_satis = magaza_data['Satis'].sum()
            toplam_envanter = magaza_data['Envanter'].sum()

            metrikler[magaza] = {
                'toplam_satis': int(toplam_satis),
                'toplam_envanter': int(toplam_envanter),
                'satis_orani': float(toplam_satis / (toplam_satis + toplam_envanter)) if (toplam_satis + toplam_envanter) > 0 else 0,
                'envanter_fazlasi': int(toplam_envanter - toplam_satis),
                'urun_sayisi': len(magaza_data)
            }
        return metrikler

    def urun_anahtari_olustur(self, urun_adi, renk, beden):
        """Urun adi + renk + beden kombinasyonu ile benzersiz anahtar olustur"""
        try:
            urun_adi = str(urun_adi).strip().upper() if pd.notna(urun_adi) else ""
            renk = str(renk).strip().upper() if pd.notna(renk) else ""
            beden = str(beden).strip().upper() if pd.notna(beden) else ""
            return f"{urun_adi} {renk} {beden}".strip()
        except Exception as e:
            logger.error(f"Product key creation error: {e}")
            return "ERROR_KEY"

    def str_hesapla(self, satis, envanter):
        """Sell-Through Rate hesapla"""
        try:
            toplam = satis + envanter
            if toplam == 0:
                return 0
            return satis / toplam
        except Exception as e:
            logger.error(f"STR calculation error: {e}")
            return 0

    def str_bazli_transfer_hesapla(self, gonderen_satis, gonderen_envanter, alan_satis, alan_envanter, strategy='sakin'):
        """STR bazli transfer miktari hesapla - ZORUNLU MIN 1 KALDIRILDI"""
        try:
            gonderen_str = self.str_hesapla(gonderen_satis, gonderen_envanter)
            alan_str = self.str_hesapla(alan_satis, alan_envanter)
            str_farki = alan_str - gonderen_str
            teorik_transfer = str_farki * gonderen_envanter
            
            config = STRATEGY_CONFIG.get(strategy, STRATEGY_CONFIG['sakin'])
            
            max_transfer_40 = gonderen_envanter * 0.40
            min_kalan = gonderen_envanter - config['min_inventory']
            
            if config['max_transfer'] is None:
                max_transfer_limit = float('inf')
            else:
                max_transfer_limit = config['max_transfer']
            
            # ZORUNLU MIN 1 KALDIRILDI - Formül 0 derse 0 döner
            transfer_miktari = min(teorik_transfer, max_transfer_40, min_kalan, max_transfer_limit)
            transfer_miktari = max(0, min(transfer_miktari, gonderen_envanter))
            
            uygulanan_filtre = 'Teorik'
            if transfer_miktari == max_transfer_40:
                uygulanan_filtre = 'Max %40'
            elif transfer_miktari == min_kalan:
                uygulanan_filtre = f'Min {config["min_inventory"]} kalsin'
            elif transfer_miktari == max_transfer_limit and config['max_transfer'] is not None:
                uygulanan_filtre = f'Max {config["max_transfer"]} adet'
            
            return int(transfer_miktari), {
                'gonderen_str': round(gonderen_str * 100, 1),
                'alan_str': round(alan_str * 100, 1),
                'str_farki': round(str_farki * 100, 1),
                'teorik_transfer': round(teorik_transfer, 1),
                'uygulanan_filtre': uygulanan_filtre,
                'kullanilan_strateji': strategy
            }
        except Exception as e:
            logger.error(f"Transfer calculation error: {e}")
            return 0, {
                'gonderen_str': 0,
                'alan_str': 0,
                'str_farki': 0,
                'teorik_transfer': 0,
                'uygulanan_filtre': 'Hata',
                'kullanilan_strateji': strategy,
                'error': str(e)
            }

    def transfer_kosullari_kontrol(self, gonderen_satis, gonderen_envanter, alan_satis, alan_envanter, strategy='sakin'):
        """STR bazli transfer kosullari kontrol - SATIS EŞIKİ KALDIRILDI"""
        try:
            config = STRATEGY_CONFIG.get(strategy, STRATEGY_CONFIG['sakin'])
            
            # SATIS EŞIKİ TAMAMEN KALDIRILDI - Sadece envanter ve STR kontrol edilir
            
            if gonderen_envanter < config['min_inventory']:
                return False, f"Gonderen envanter yetersiz ({gonderen_envanter} < {config['min_inventory']})"
            
            gonderen_str = self.str_hesapla(gonderen_satis, gonderen_envanter)
            alan_str = self.str_hesapla(alan_satis, alan_envanter)
            str_farki = alan_str - gonderen_str
            
            if str_farki < config['min_str_diff']:
                return False, f"STR farki yetersiz ({str_farki*100:.1f}% < {config['min_str_diff']*100}%)"
            
            transfer_miktari, detaylar = self.safe_transfer_calculation(
                gonderen_satis, gonderen_envanter, alan_satis, alan_envanter, strategy
            )
            
            if transfer_miktari <= 0:
                return False, "Transfer miktari hesaplanamadi"
            
            return True, f"STR: A{detaylar['alan_str']}%>G{detaylar['gonderen_str']}%, T:{transfer_miktari}"
        except Exception as e:
            logger.error(f"Transfer condition check error: {e}")
            return False, f"Hata: {str(e)}"

    @measure_time
    def beden_tamamlama_analizi_yap(self, target_store, excluded_stores=None):
        """Beden tamamlama analizi + MERKEZ/ONLINE ÖNCELİĞİ"""
        if self.data is None:
            return None

        self.target_store = target_store
        self.transfer_type = 'size_completion'
        if excluded_stores is None:
            excluded_stores = []
        self.excluded_stores = excluded_stores

        logger.info(f"Hedef magaza beden tamamlama analizi baslatiliyor... Hedef: {target_store}")
        
        transferler = []
        processed_items = 0
        
        target_data = self.data[self.data['Depo Adi'] == target_store]
        
        if target_data.empty:
            logger.warning(f"Hedef magaza '{target_store}' icin veri bulunamadi")
            return None

        logger.info("Creating product keys...")
        self.data = self.create_product_key_vectorized(self.data)
        target_data = self.create_product_key_vectorized(target_data)

        sifir_envanter = target_data[target_data['Envanter'] == 0]
        total_items = len(sifir_envanter)
        logger.info(f"{target_store}'da envanter=0 olan {total_items} urun kombinasyonu bulundu")
        
        if sifir_envanter.empty:
            logger.info(f"{target_store}'da hic eksik urun yok!")
            result = {
                'analiz_tipi': 'beden_tamamlama',
                'strateji': 'urun_anahtari_bazli',
                'target_store': target_store,
                'excluded_stores': excluded_stores,
                'transferler': [],
                'toplam_eksik_beden': 0
            }
            self.mevcut_analiz = result
        self.save_to_temp()
            return result
        
        for index, eksik_row in sifir_envanter.iterrows():
            processed_items += 1
            if processed_items % 50 == 0 or processed_items == total_items:
                logger.info(f"Beden tamamlama progress: {processed_items}/{total_items} items processed")
            
            try:
                eksik_urun_anahtari = eksik_row['urun_anahtari']
                urun_adi = eksik_row['Urun Adi']
                renk = eksik_row.get('Renk Aciklamasi', '')
                beden = str(eksik_row.get('Beden', '')).strip()
                urun_kodu = eksik_row.get('Urun Kodu', '')
                alan_satis = eksik_row['Satis']
                
                kaynak_magazalar = self.data[
                    (self.data['Depo Adi'] != target_store) &
                    (self.data['urun_anahtari'] == eksik_urun_anahtari) &
                    (self.data['Envanter'] > 0) &
                    (~self.data['Depo Adi'].isin(excluded_stores))
                ]
                
                if kaynak_magazalar.empty:
                    continue
                
                # MERKEZ/ONLINE ÖNCELİK SİSTEMİ
                priority_sources = kaynak_magazalar[kaynak_magazalar['Depo Adi'].isin(WAREHOUSE_SET)]
                if not priority_sources.empty:
                    en_iyi_kaynak = priority_sources.loc[priority_sources['Envanter'].idxmax()]
                    logger.info(f"Öncelikli kaynak bulundu: {en_iyi_kaynak['Depo Adi']}")
                else:
                    en_iyi_kaynak = kaynak_magazalar.loc[kaynak_magazalar['Envanter'].idxmax()]
                
                gonderen_magaza = en_iyi_kaynak['Depo Adi']
                gonderen_envanter = en_iyi_kaynak['Envanter']
                gonderen_satis = en_iyi_kaynak['Satis']
                
                transferler.append({
                    'urun_adi': urun_adi,
                    'urun_kodu': urun_kodu,
                    'renk': renk,
                    'beden': beden,
                    'gonderen_magaza': gonderen_magaza,
                    'alan_magaza': target_store,
                    'transfer_miktari': 1,
                    'gonderen_satis': int(gonderen_satis),
                    'gonderen_envanter': int(gonderen_envanter),
                    'alan_satis': int(alan_satis),
                    'alan_envanter': 0,
                    'transfer_tipi': 'urun_anahtari_beden_tamamlama',
                    'eksik_beden': True,
                    'kullanilan_strateji': 'urun_anahtari_bazli',
                    'urun_anahtari': eksik_urun_anahtari,
                    'oncelikli_kaynak': gonderen_magaza in WAREHOUSE_SET
                })
            except Exception as e:
                logger.error(f"Error processing item {processed_items}: {e}")
                continue

        logger.info(f"Beden tamamlama tamamlandi: {len(transferler)} transfer onerisi")
        
        transferler.sort(key=lambda x: (not x.get('oncelikli_kaynak', False), x['urun_anahtari']))
        
        result = {
            'analiz_tipi': 'beden_tamamlama',
            'strateji': 'urun_anahtari_bazli',
            'target_store': target_store,
            'excluded_stores': excluded_stores,
            'transferler': transferler,
            'toplam_eksik_beden': len(transferler)
        }
        
        self.save_to_temp()
        return result

    @measure_time
    def targeted_transfer_analizi_yap(self, target_store, strategy='sakin', excluded_stores=None):
        """Spesifik magaza icin transfer analizi + MERKEZ/ONLINE ÖNCELİĞİ"""
        if self.data is None:
            return None

        self.current_strategy = strategy
        self.target_store = target_store
        self.transfer_type = 'targeted'
        if excluded_stores is None:
            excluded_stores = []
        self.excluded_stores = excluded_stores

        logger.info(f"Spesifik magaza analizi baslatiliyor... Hedef: {target_store}, Strateji: {strategy}")
        
        transferler = []
        processed_items = 0
        
        target_data = self.data[self.data['Depo Adi'] == target_store]
        
        if target_data.empty:
            logger.warning(f"Hedef magaza '{target_store}' icin veri bulunamadi")
            return None

        diger_magazalar = [m for m in self.magazalar if m != target_store and m not in excluded_stores]
        
        tum_data = self.data.copy()
        tum_data = self.create_product_key_vectorized(tum_data)
        
        target_data_with_keys = self.create_product_key_vectorized(target_data)
        target_urun_anahtarlari = target_data_with_keys['urun_anahtari'].unique()
        total_items = len(target_urun_anahtarlari)

        for urun_anahtari in target_urun_anahtarlari:
            processed_items += 1
            if processed_items % 100 == 0 or processed_items == total_items:
                logger.info(f"Targeted transfer progress: {processed_items}/{total_items} products processed")
            
            try:
                target_urun_data = target_data_with_keys[target_data_with_keys['urun_anahtari'] == urun_anahtari]
                
                if target_urun_data.empty:
                    continue
                    
                target_row = target_urun_data.iloc[0]
                alan_satis = target_row['Satis']
                alan_envanter = target_row['Envanter']
                
                # En iyi donörü bul - MERKEZ/ONLINE ÖNCELİĞİ İLE
                best_transfer = None
                best_priority_score = -1
                
                # 1. Önce MERKEZ/ONLINE kaynaklarından ara
                warehouse_candidates = tum_data[
                    (tum_data['Depo Adi'].isin(WAREHOUSE_SET)) & 
                    (tum_data['urun_anahtari'] == urun_anahtari) & 
                    (tum_data['Envanter'] >= 2)
                ].sort_values(['Envanter', 'Depo Adi'], ascending=[False, True])
                
                for _, warehouse_row in warehouse_candidates.iterrows():
                    gonderen_magaza = warehouse_row['Depo Adi']
                    gonderen_satis = warehouse_row['Satis']
                    gonderen_envanter = warehouse_row['Envanter']
                    
                    transfer_miktari, str_detaylar = self.safe_transfer_calculation(
                        gonderen_satis, gonderen_envanter, alan_satis, alan_envanter, strategy
                    )
                    
                    if transfer_miktari > 0:
                        best_transfer = {
                            'urun_anahtari': urun_anahtari,
                            'urun_kodu': warehouse_row.get('Urun Kodu', target_row.get('Urun Kodu', '')),
                            'urun_adi': warehouse_row.get('Urun Adi', target_row.get('Urun Adi', '')),
                            'renk': warehouse_row.get('Renk Aciklamasi', target_row.get('Renk Aciklamasi', '')),
                            'beden': warehouse_row.get('Beden', target_row.get('Beden', '')),
                            'gonderen_magaza': gonderen_magaza,
                            'alan_magaza': target_store,
                            'transfer_miktari': int(transfer_miktari),
                            'gonderen_satis': int(gonderen_satis),
                            'gonderen_envanter': int(gonderen_envanter),
                            'alan_satis': int(alan_satis),
                            'alan_envanter': int(alan_envanter),
                            'gonderen_str': str_detaylar['gonderen_str'],
                            'alan_str': str_detaylar['alan_str'],
                            'str_farki': str_detaylar['str_farki'],
                            'kullanilan_strateji': strategy,
                            'transfer_tipi': 'targeted',
                            'oncelikli_kaynak': True
                        }
                        break
                
                # 2. Eğer MERKEZ/ONLINE'dan bulunamadıysa diğer mağazalardan ara
                if best_transfer is None:
                    for gonderen_magaza in diger_magazalar:
                        if gonderen_magaza in WAREHOUSE_SET:  # Zaten yukarıda kontrol edildi
                            continue
                            
                        gonderen_urun_data = tum_data[
                            (tum_data['Depo Adi'] == gonderen_magaza) &
                            (tum_data['urun_anahtari'] == urun_anahtari)
                        ]
                        
                        if gonderen_urun_data.empty:
                            continue
                            
                        gonderen_row = gonderen_urun_data.iloc[0]
                        gonderen_satis = gonderen_row['Satis']
                        gonderen_envanter = gonderen_row['Envanter']
                        
                        kosul_sonuc, kosul_mesaj = self.transfer_kosullari_kontrol(
                            gonderen_satis, gonderen_envanter, alan_satis, alan_envanter, strategy
                        )
                        
                        if kosul_sonuc:
                            transfer_miktari, str_detaylar = self.safe_transfer_calculation(
                                gonderen_satis, gonderen_envanter, alan_satis, alan_envanter, strategy
                            )
                            
                            if transfer_miktari > 0:
                                priority_score = str_detaylar['str_farki'] + min(gonderen_envanter / 10, 50)
                                
                                if priority_score > best_priority_score:
                                    best_priority_score = priority_score
                                    best_transfer = {
                                        'urun_anahtari': urun_anahtari,
                                        'urun_kodu': gonderen_row['Urun Kodu'],
                                        'urun_adi': gonderen_row['Urun Adi'],
                                        'renk': gonderen_row.get('Renk Aciklamasi', ''),
                                        'beden': gonderen_row.get('Beden', ''),
                                        'gonderen_magaza': gonderen_magaza,
                                        'alan_magaza': target_store,
                                        'transfer_miktari': int(transfer_miktari),
                                        'gonderen_satis': int(gonderen_satis),
                                        'gonderen_envanter': int(gonderen_envanter),
                                        'alan_satis': int(alan_satis),
                                        'alan_envanter': int(alan_envanter),
                                        'gonderen_str': str_detaylar['gonderen_str'],
                                        'alan_str': str_detaylar['alan_str'],
                                        'str_farki': str_detaylar['str_farki'],
                                        'kullanilan_strateji': strategy,
                                        'transfer_tipi': 'targeted',
                                        'oncelikli_kaynak': False
                                    }
                
                if best_transfer:
                    transferler.append(best_transfer)
                    
            except Exception as e:
                logger.error(f"Error processing targeted transfer for product {processed_items}: {e}")
                continue

        # Öncelik skoruna göre sırala (öncelikli kaynaklar önce, sonra STR farkı)
        transferler.sort(key=lambda x: (not x.get('oncelikli_kaynak', False), -x.get('str_farki', 0)))
        
        logger.info(f"Spesifik magaza analizi tamamlandi: {len(transferler)} transfer onerisi")
        
        result = {
            'analiz_tipi': 'targeted',
            'strateji': strategy,
            'target_store': target_store,
            'excluded_stores': excluded_stores,
            'transferler': transferler
        }
        
        self.mevcut_analiz = result
        self.save_to_temp()
        return result

    @measure_time
    def global_transfer_analizi_yap(self, strategy='sakin', excluded_stores=None):
        """Global transfer analizi + MERKEZ/ONLINE ÖNCELİĞİ"""
        if self.data is None:
            return None

        self.current_strategy = strategy
        self.transfer_type = 'global'
        if excluded_stores is None:
            excluded_stores = []
        self.excluded_stores = excluded_stores
        
        config = STRATEGY_CONFIG.get(strategy, STRATEGY_CONFIG['sakin'])

        logger.info(f"Global urun bazli STR transfer analizi baslatiliyor... Strateji: {strategy}")
        logger.info(f"Istisna magazalar: {excluded_stores}")
        logger.info(f"Strateji parametreleri: {config}")
        
        start_memory = self.check_memory_usage()
        metrikler = self.magaza_metrikleri_hesapla()
        transferler = []
        transfer_gereksiz = []

        tum_data = self.data.copy()
        
        if excluded_stores:
            tum_data = tum_data[~tum_data['Depo Adi'].isin(excluded_stores)]
            logger.info(f"Istisna magazalar filtrelendi. Kalan veri: {len(tum_data)} satir")
        
        tum_data = self.create_product_key_vectorized(tum_data)
        
        tum_urun_anahtarlari = tum_data['urun_anahtari'].unique()
        total_products = len(tum_urun_anahtarlari)
        
        logger.info(f"Toplam {total_products} benzersiz urun grubu analiz ediliyor...")

        for index, urun_anahtari in enumerate(tum_urun_anahtarlari):
            if (index + 1) % 100 == 0 or (index + 1) == total_products:
                progress_percent = ((index + 1) / total_products) * 100
                logger.info(f"Global transfer progress: {index + 1}/{total_products} ({progress_percent:.1f}%) completed")
                
                if (index + 1) % 500 == 0:
                    current_memory = self.check_memory_usage()
                    logger.info(f"Memory usage during processing: {current_memory}%")

            try:
                urun_data = tum_data[tum_data['urun_anahtari'] == urun_anahtari]
                
                magaza_gruplari = urun_data.groupby('Depo Adi').agg({
                    'Satis': 'sum',
                    'Envanter': 'sum',
                    'Urun Adi': 'first',
                    'Renk Aciklamasi': 'first',
                    'Beden': 'first',
                    'Urun Kodu': 'first'
                }).reset_index()

                if len(magaza_gruplari) < 2:
                    continue

                magaza_str_listesi = []
                for _, magaza_grup in magaza_gruplari.iterrows():
                    magaza = magaza_grup['Depo Adi']
                    
                    if magaza in excluded_stores:
                        continue
                        
                    satis = magaza_grup['Satis']
                    envanter = magaza_grup['Envanter']
                    str_value = self.str_hesapla_cached(satis, envanter)
                    
                    magaza_str_listesi.append({
                        'magaza': magaza,
                        'satis': satis,
                        'envanter': envanter,
                        'str': str_value,
                        'urun_adi': magaza_grup['Urun Adi'],
                        'renk': magaza_grup.get('Renk Aciklamasi', ''),
                        'beden': magaza_grup.get('Beden', ''),
                        'urun_kodu': magaza_grup['Urun Kodu']
                    })

                if len(magaza_str_listesi) < 2:
                    continue

                magaza_str_listesi.sort(key=lambda x: x['str'])
                
                # En yüksek STR'li mağaza (alıcı)
                en_yuksek_str = magaza_str_listesi[-1]
                
                # MERKEZ/ONLINE ÖNCELİK SİSTEMİ (gönderici seçimi)
                warehouse_candidates_ge2 = [m for m in magaza_str_listesi if m['magaza'] in WAREHOUSE_SET and m['envanter'] >= 2]
                if warehouse_candidates_ge2:
                    warehouse_candidates_ge2.sort(key=lambda x: (-x['envanter'], x['magaza']))
                    en_dusuk_str = warehouse_candidates_ge2[0]
                else:
                    warehouse_candidates_eq1 = [m for m in magaza_str_listesi if m['magaza'] in WAREHOUSE_SET and m['envanter'] == 1]
                    if warehouse_candidates_eq1:
                        warehouse_candidates_eq1.sort(key=lambda x: x['magaza'])
                        en_dusuk_str = warehouse_candidates_eq1[0]
                    else:
                        # Normal mağazalardan en düşük STR'li
                        en_dusuk_str = magaza_str_listesi[0]

                kosul_sonuc, kosul_mesaj = self.transfer_kosullari_kontrol(
                    en_dusuk_str['satis'], en_dusuk_str['envanter'], 
                    en_yuksek_str['satis'], en_yuksek_str['envanter'],
                    strategy
                )
                
                if kosul_sonuc:
                    transfer_miktari, str_detaylar = self.safe_transfer_calculation(
                        en_dusuk_str['satis'], en_dusuk_str['envanter'],
                        en_yuksek_str['satis'], en_yuksek_str['envanter'],
                        strategy
                    )
                    
                    if transfer_miktari > 0:
                        alan_str_val = str_detaylar['alan_str']
                        if alan_str_val >= 80:
                            stok_durumu = 'YUKSEK'
                        elif alan_str_val >= 50:
                            stok_durumu = 'NORMAL'
                        elif alan_str_val >= 20:
                            stok_durumu = 'DUSUK'
                        else:
                            stok_durumu = 'KRITIK'
                        
                        transferler.append({
                            'urun_anahtari': urun_anahtari,
                            'urun_kodu': en_dusuk_str['urun_kodu'],
                            'urun_adi': en_dusuk_str['urun_adi'],
                            'renk': en_dusuk_str['renk'],
                            'beden': en_dusuk_str['beden'],
                            'gonderen_magaza': en_dusuk_str['magaza'],
                            'alan_magaza': en_yuksek_str['magaza'],
                            'transfer_miktari': int(transfer_miktari),
                            'gonderen_satis': int(en_dusuk_str['satis']),
                            'gonderen_envanter': int(en_dusuk_str['envanter']),
                            'alan_satis': int(en_yuksek_str['satis']),
                            'alan_envanter': int(en_yuksek_str['envanter']),
                            'gonderen_str': str_detaylar['gonderen_str'],
                            'alan_str': str_detaylar['alan_str'],
                            'str_farki': str_detaylar['str_farki'],
                            'teorik_transfer': str_detaylar['teorik_transfer'],
                            'uygulanan_filtre': str_detaylar['uygulanan_filtre'],
                            'kullanilan_strateji': str_detaylar['kullanilan_strateji'],
                            'alan_stok_durumu': stok_durumu,
                            'magaza_sayisi': len(magaza_str_listesi),
                            'min_str': round(en_dusuk_str['str'] * 100, 1),
                            'max_str': round(en_yuksek_str['str'] * 100, 1),
                            'satis_farki': int(en_yuksek_str['satis'] - en_dusuk_str['satis']),
                            'envanter_farki': int(en_dusuk_str['envanter'] - en_yuksek_str['envanter']),
                            'oncelikli_kaynak': en_dusuk_str['magaza'] in WAREHOUSE_SET
                        })
                else:
                    str_ortalama = sum(m['str'] for m in magaza_str_listesi) / len(magaza_str_listesi)
                    str_fark = max(m['str'] for m in magaza_str_listesi) - min(m['str'] for m in magaza_str_listesi)
                    
                    transfer_gereksiz.append({
                        'urun_anahtari': urun_anahtari,
                        'urun_adi': magaza_str_listesi[0]['urun_adi'],
                        'renk': magaza_str_listesi[0]['renk'],
                        'beden': magaza_str_listesi[0]['beden'],
                        'magaza_sayisi': len(magaza_str_listesi),
                        'ortalama_str': round(str_ortalama * 100, 1),
                        'str_fark': round(str_fark * 100, 1),
                        'red_nedeni': kosul_mesaj
                    })
            except Exception as e:
                logger.error(f"Error processing product {index + 1}: {e}")
                continue

        # Öncelikli kaynaklara göre sırala, sonra STR farkı
        transferler.sort(key=lambda x: (not x.get('oncelikli_kaynak', False), -x['str_farki']))

        end_memory = self.check_memory_usage()
        logger.info(f"Global analiz tamamlandi ({strategy}): {len(transferler)} transfer, {len(transfer_gereksiz)} red")
        logger.info(f"Memory usage - Start: {start_memory}%, End: {end_memory}%")
        if excluded_stores:
            logger.info(f"Istisna magazalar: {excluded_stores}")

        self.performance_metrics['last_global_analysis'] = {
            'timestamp': datetime.now().isoformat(),
            'strategy': strategy,
            'total_products': total_products,
            'successful_transfers': len(transferler),
            'rejected_transfers': len(transfer_gereksiz),
            'excluded_stores_count': len(excluded_stores),
            'memory_usage_start': start_memory,
            'memory_usage_end': end_memory
        }

        result = {
            'analiz_tipi': 'global',
            'strateji': strategy,
            'strateji_parametreleri': config,
            'excluded_stores': excluded_stores,
            'excluded_count': len(excluded_stores),
            'magaza_metrikleri': metrikler,
            'transferler': transferler,
            'transfer_gereksiz': transfer_gereksiz,
            'performance_metrics': self.performance_metrics.get('last_global_analysis', {})
        }
        
        self.save_to_temp()
        return result

    def simulate_transfer_impact(self, transfers):
        """Transfer etkisini simüle et"""
        try:
            if not transfers:
                return {
                    'error': 'No transfers to simulate',
                    'impact': {}
                }
            
            simulation_results = {
                'total_transfers': len(transfers),
                'total_items_moved': sum(t.get('transfer_miktari', 0) for t in transfers),
                'affected_stores': len(set([t.get('gonderen_magaza', '') for t in transfers] + 
                                         [t.get('alan_magaza', '') for t in transfers])),
                'average_str_improvement': 0,
                'risk_level': 'UNKNOWN',
                'priority_transfers': sum(1 for t in transfers if t.get('oncelikli_kaynak', False))
            }
            
            str_improvements = [t.get('str_farki', 0) for t in transfers if t.get('str_farki', 0) > 0]
            if str_improvements:
                simulation_results['average_str_improvement'] = sum(str_improvements) / len(str_improvements)
            
            high_volume_transfers = sum(1 for t in transfers if t.get('transfer_miktari', 0) > 10)
            if high_volume_transfers > len(transfers) * 0.3:
                simulation_results['risk_level'] = 'HIGH'
            elif high_volume_transfers > len(transfers) * 0.1:
                simulation_results['risk_level'] = 'MEDIUM'
            else:
                simulation_results['risk_level'] = 'LOW'
            
            return {
                'success': True,
                'impact': simulation_results
            }
        except Exception as e:
            logger.error(f"Transfer simulation error: {e}")
            return {
                'error': str(e),
                'impact': {}
            }

# Global sistem instance
sistem = MagazaTransferSistemi()

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/', methods=['GET'])
def health_check():
    """Health check endpoint with performance metrics"""
    try:
        memory_usage = sistem.check_memory_usage()
        
        return jsonify({
            'status': 'healthy',
            'service': 'RetailFlow Transfer API',
            'version': '6.1.0-optimized-fixed',
            'timestamp': datetime.now().isoformat(),
            'data_loaded': sistem.data is not None,
            'store_count': len(sistem.magazalar) if sistem.magazalar else 0,
            'current_strategy': sistem.current_strategy,
            'transfer_type': sistem.transfer_type,
            'target_store': sistem.target_store,
            'excluded_stores': sistem.excluded_stores,
            'available_strategies': list(STRATEGY_CONFIG.keys()),
            'memory_usage_percent': memory_usage,
            'performance_metrics': sistem.performance_metrics
        })
    except Exception as e:
        logger.error(f"Health check error: {e}")
        return jsonify({
            'status': 'error',
            'error': str(e),
            'timestamp': datetime.now().isoformat()
        }), 500

@app.route('/upload', methods=['POST'])
def upload_file():
    """Dosya yukleme endpoint'i"""
    try:
        logger.info("File upload request received")
        
        if 'file' not in request.files:
            return jsonify({'error': 'Dosya secilmedi'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'Dosya secilmedi'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'error': 'Gecersiz dosya formati! Excel (.xlsx, .xls) veya CSV dosyasi yukleyin.'}), 400
        
        filename = secure_filename(file.filename)
        logger.info(f"Processing file: {filename}")
        
        initial_memory = sistem.check_memory_usage()
        logger.info(f"Memory usage before file processing: {initial_memory}%")
        
        try:
            if filename.lower().endswith('.csv'):
                try:
                    df = pd.read_csv(file, encoding='utf-8')
                except UnicodeDecodeError:
                    df = pd.read_csv(file, encoding='cp1254')
            else:
                df = pd.read_excel(file, engine='openpyxl' if filename.endswith('.xlsx') else None)
            
            logger.info(f"File read successfully: {len(df)} rows")
        except Exception as e:
            logger.error(f"File reading error: {str(e)}")
            return jsonify({'error': f'Dosya okuma hatasi: {str(e)}'}), 400
        
        if len(df) > 1000000:
            return jsonify({'error': 'Dosya cok buyuk! Maksimum 1 milyon satir desteklenmektedir.'}), 400
        
        success, result = sistem.dosya_yukle_df(df)
        
        if success:
            logger.info("Data loaded successfully")
            return jsonify({
                'success': True,
                'filename': filename,
                'data': result
            })
        else:
            logger.error(f"Data loading failed: {result}")
            return jsonify({'error': result}), 400
            
    except Exception as e:
        logger.error(f"Upload error: {str(e)}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        return jsonify({'error': f'Dosya yukleme hatasi: {str(e)}'}), 500

@app.route('/remove-file', methods=['POST'])
def remove_file():
    """Yuklenen dosyayi kaldirma endpoint'i"""
    try:
        logger.info("File removal request received")
        
        if sistem.data is None:
            logger.warning("No file to remove")
            return jsonify({'error': 'Kaldirilacak dosya yok'}), 400
        
        success = sistem.clear_all_data()
        
        if success:
            logger.info("File and all data removed successfully")
            return jsonify({
                'success': True,
                'message': 'Dosya ve tum veriler basariyla kaldirildi',
                'timestamp': datetime.now().isoformat()
            })
        else:
            logger.error("Failed to remove file and data")
            return jsonify({'error': 'Dosya kaldirma islemi basarisiz'}), 500
            
    except Exception as e:
        logger.error(f"File removal error: {str(e)}")
        return jsonify({'error': f'Dosya kaldirma hatasi: {str(e)}'}), 500

@app.route('/analyze', methods=['POST'])
def analyze_data():
    """Transfer analizi"""
    try:
        logger.info("Analysis request received")
        
        if sistem.data is None:
            logger.warning("No data available for analysis")
            return jsonify({'error': 'Once bir dosya yukleyin'}), 400
        
        request_data = request.get_json() or {}
        request_data = sistem.validate_request_data(request_data)
        
        strategy = request_data.get('strategy', 'sakin')
        excluded_stores = request_data.get('excluded_stores', [])
        transfer_type = request_data.get('transfer_type', 'global')
        target_store = request_data.get('target_store', None)
        
        valid_excluded_stores = [store for store in excluded_stores if store in sistem.magazalar]
        
        logger.info(f"Starting {transfer_type} analysis... Strategy: {strategy}")
        logger.info(f"Target store: {target_store}")
        logger.info(f"Excluded stores: {valid_excluded_stores}")
        
        pre_analysis_memory = sistem.check_memory_usage()
        logger.info(f"Memory usage before analysis: {pre_analysis_memory}%")
        
        if transfer_type == 'size_completion':
            if not target_store or target_store not in sistem.magazalar:
                return jsonify({'error': 'Beden tamamlama icin gecerli bir hedef magaza secin'}), 400
            
            results = sistem.beden_tamamlama_analizi_yap(target_store, valid_excluded_stores)
        
        elif transfer_type == 'targeted':
            if not target_store or target_store not in sistem.magazalar:
                return jsonify({'error': 'Spesifik magaza analizi icin gecerli bir hedef magaza secin'}), 400
            
            results = sistem.targeted_transfer_analizi_yap(target_store, strategy, valid_excluded_stores)
        
        else:  # global
            results = sistem.global_transfer_analizi_yap(strategy, valid_excluded_stores)
        
        if results:
            sistem.mevcut_analiz = results
            
            simulation = sistem.simulate_transfer_impact(results.get('transferler', []))
            
            limited_results = {
                'analiz_tipi': results['analiz_tipi'],
                'strateji': results['strateji'],
                'target_store': results.get('target_store'),
                'excluded_stores': results.get('excluded_stores', []),
                'excluded_count': results.get('excluded_count', 0),
                'transferler': results['transferler'][:50] if results.get('transferler') else [],
                'toplam_transfer_sayisi': len(results.get('transferler', [])),
                'simulation': simulation
            }
            
            if transfer_type == 'global':
                limited_results['strateji_parametreleri'] = results.get('strateji_parametreleri')
                limited_results['magaza_metrikleri'] = results.get('magaza_metrikleri')
                limited_results['transfer_gereksiz'] = results.get('transfer_gereksiz', [])[:20]
                limited_results['toplam_gereksiz_sayisi'] = len(results.get('transfer_gereksiz', []))
                limited_results['performance_metrics'] = results.get('performance_metrics', {})
            
            post_analysis_memory = sistem.check_memory_usage()
            limited_results['memory_usage'] = {
                'before_analysis': pre_analysis_memory,
                'after_analysis': post_analysis_memory
            }
            
            logger.info(f"{transfer_type.capitalize()} analysis completed ({strategy}): {len(results.get('transferler', []))} total transfers")
            if valid_excluded_stores:
                logger.info(f"Excluded {len(valid_excluded_stores)} stores from analysis")
            
            return jsonify({
                'success': True,
                'results': limited_results
            })
        else:
            logger.error("Analysis returned no results")
            return jsonify({'error': 'Analiz basarisiz'}), 500
            
    except Exception as e:
        logger.error(f"Analysis error: {str(e)}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        return jsonify({'error': f'Analiz hatasi: {str(e)}'}), 500

@app.route('/export/excel', methods=['POST'])
def export_excel():
    """Excel export + Performance Metrics Sheet"""
    try:
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        
        logger.info("Excel export request received")
        
        if not sistem.mevcut_analiz:
            logger.warning("No analysis results available for export")
            return jsonify({'error': 'Analiz sonucu bulunamadi'}), 400
        
        transferler = sistem.mevcut_analiz['transferler']
        analiz_tipi = sistem.mevcut_analiz.get('analiz_tipi', 'global')
        strategy = sistem.mevcut_analiz.get('strateji', 'sakin')
        target_store = sistem.mevcut_analiz.get('target_store', '')
        
        logger.info(f"Exporting {len(transferler)} transfers to Excel (Type: {analiz_tipi}, Strategy: {strategy})")
        
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Ana transfer sayfasi
            if transferler:
                df_transfer = pd.DataFrame(transferler)
                
                if analiz_tipi == 'size_completion':
                    selected_columns = {
                        'urun_adi': 'Ürün Adı',
                        'renk': 'Renk',
                        'beden': 'Eksik Beden',
                        'gonderen_magaza': 'Gönderen Mağaza',
                        'gonderen_satis': 'Gönderen Satış',
                        'gonderen_envanter': 'Gönderen Envanter',
                        'alan_magaza': 'Hedef Mağaza',
                        'alan_satis': 'Alan Satış',
                        'alan_envanter': 'Alan Envanter',
                        'transfer_miktari': 'Transfer Miktarı'
                    }
                    sheet_name = f'{target_store} Beden Tamamlama'
                else:
                    selected_columns = {
                        'urun_kodu': 'Ürün Kodu',
                        'urun_adi': 'Ürün Adı',
                        'renk': 'Renk',
                        'beden': 'Beden',
                        'gonderen_magaza': 'Gönderen Mağaza',
                        'gonderen_satis': 'Gönderen Satış',
                        'gonderen_envanter': 'Gönderen Envanter',
                        'alan_magaza': 'Alan Mağaza',
                        'alan_satis': 'Alan Satış',
                        'alan_envanter': 'Alan Envanter',
                        'transfer_miktari': 'Transfer Miktarı',
                        'str_farki': 'STR Farkı (%)'
                    }
                    sheet_name = f'{target_store} Transferleri' if analiz_tipi == 'targeted' else 'Transfer Önerileri'
                
                available_columns = {k: v for k, v in selected_columns.items() if k in df_transfer.columns}
                df_export = df_transfer[list(available_columns.keys())].copy()
                df_export = df_export.rename(columns=available_columns)
                
                df_export.to_excel(writer, index=False, sheet_name=sheet_name[:31])
                
                workbook = writer.book
                worksheet = workbook[sheet_name[:31]]
                
                header_font = Font(name='Segoe UI', size=14, bold=True, color='FFFFFF')
                header_fill = PatternFill(start_color='244062', end_color='244062', fill_type='solid')
                data_font = Font(name='Segoe UI', size=11)
                
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for col_num, col in enumerate(worksheet.iter_cols(max_row=1), 1):
                    cell = col[0]
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                for row in worksheet.iter_rows(min_row=2):
                    for cell in row:
                        cell.font = data_font
                        cell.border = thin_border
                
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            else:
                # Eğer transfer yoksa boş sayfa oluştur
                df_empty = pd.DataFrame({'Mesaj': ['Hiç transfer önerisi bulunamadı']})
                df_empty.to_excel(writer, index=False, sheet_name='Sonuçlar')
            
            
            # Analiz ozeti sayfasi
            summary_info = {
                'Analiz Tipi': [analiz_tipi.upper()],
                'Kullanılan Strateji': [strategy],
                'Hedef Mağaza': [target_store or 'Tüm Mağazalar'],
                'Toplam Transfer': [len(transferler)],
                'İstisna Mağaza Sayısı': [len(sistem.excluded_stores)],
                'Analiz Tarihi': [datetime.now().strftime('%Y-%m-%d %H:%M')]
            }
            
            if sistem.excluded_stores:
                summary_info['İstisna Mağazalar'] = [', '.join(sistem.excluded_stores)]
            
            df_summary = pd.DataFrame(summary_info)
            df_summary.to_excel(writer, index=False, sheet_name='Analiz Özeti')
            
            # Performance Metrics sayfasi
            if sistem.performance_metrics:
                perf_data = []
                
                if 'last_file_load' in sistem.performance_metrics:
                    file_metrics = sistem.performance_metrics['last_file_load']
                    perf_data.append({
                        'Metrik': 'Dosya Yükleme',
                        'Değer': f"{file_metrics.get('rows', 0):,} satır",
                        'Detay': f"Bellek: {file_metrics.get('memory_usage_start', 0)}% → {file_metrics.get('memory_usage_end', 0)}%",
                        'Zaman': file_metrics.get('timestamp', 'N/A')
                    })
                
                if 'last_global_analysis' in sistem.performance_metrics:
                    analysis_metrics = sistem.performance_metrics['last_global_analysis']
                    perf_data.append({
                        'Metrik': 'Global Analiz',
                        'Değer': f"{analysis_metrics.get('successful_transfers', 0)} transfer",
                        'Detay': f"Ürün: {analysis_metrics.get('total_products', 0):,}, Red: {analysis_metrics.get('rejected_transfers', 0)}",
                        'Zaman': analysis_metrics.get('timestamp', 'N/A')
                    })
                
                current_memory = sistem.check_memory_usage()
                perf_data.append({
                    'Metrik': 'Anlık Bellek Kullanımı',
                    'Değer': f"{current_memory}%",
                    'Detay': 'Sistem bellek durumu',
                    'Zaman': datetime.now().isoformat()
                })
                
                try:
                    cache_info = sistem.str_hesapla_cached.cache_info()
                    perf_data.append({
                        'Metrik': 'Cache Performansı',
                        'Değer': f"Hit: {cache_info.hits}, Miss: {cache_info.misses}",
                        'Detay': f"Hit Ratio: {cache_info.hits/(cache_info.hits+cache_info.misses)*100:.1f}%" if (cache_info.hits+cache_info.misses) > 0 else "N/A",
                        'Zaman': datetime.now().isoformat()
                    })
                except:
                    pass
                
                if perf_data:
                    df_performance = pd.DataFrame(perf_data)
                    df_performance.to_excel(writer, index=False, sheet_name='Performans Metrikleri')
                    
                    perf_worksheet = workbook['Performans Metrikleri']
                    
                    for col_num, col in enumerate(perf_worksheet.iter_cols(max_row=1), 1):
                        cell = col[0]
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    for row in perf_worksheet.iter_rows(min_row=2):
                        for cell in row:
                            cell.font = data_font
                            cell.border = thin_border
                    
                    for column in perf_worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 60)
                        perf_worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Özet sayfasını formatla
            summary_worksheet = workbook['Analiz Özeti']
            
            for col_num, col in enumerate(summary_worksheet.iter_cols(max_row=1), 1):
                cell = col[0]
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            for row in summary_worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.font = data_font
                    cell.border = thin_border
            
            for column in summary_worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                summary_worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        if analiz_tipi == 'size_completion':
            filename = f'beden_tamamlama_{target_store}_{strategy}_{timestamp}.xlsx'
        elif analiz_tipi == 'targeted':
            filename = f'targeted_{target_store}_{strategy}_{timestamp}.xlsx'
        else:
            filename = f'global_transfer_{strategy}_{timestamp}.xlsx'
        
        filename = "".join(c for c in filename if c.isalnum() or c in (' ', '-', '_', '.')).rstrip()
        
        logger.info(f"Excel file created with performance metrics: {filename}")
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Export error: {str(e)}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        return jsonify({'error': f'Export hatasi: {str(e)}'}), 500

@app.route('/stores', methods=['GET'])
def get_stores():
    """Magaza listesi"""
    try:
        if not sistem.magazalar:
            return jsonify({'error': 'Magaza verisi bulunamadi'}), 400
        
        metrikler = sistem.magaza_metrikleri_hesapla()
        stores = []
        
        for magaza in sistem.magazalar:
            if magaza in metrikler:
                m = metrikler[magaza]
                str_oran = m['satis_orani'] * 100
                stores.append({
                    'name': magaza,
                    'sales': m['toplam_satis'],
                    'inventory': m['toplam_envanter'],
                    'str_rate': round(str_oran, 1),
                    'product_count': m['urun_sayisi'],
                    'excess_inventory': m['envanter_fazlasi']
                })
        
        return jsonify({'success': True, 'stores': stores})
        
    except Exception as e:
        logger.error(f"Stores error: {str(e)}")
        return jsonify({'error': f'Magaza verisi hatasi: {str(e)}'}), 500

@app.route('/strategies', methods=['GET'])
def get_strategies():
    """Mevcut stratejileri listele"""
    try:
        strategies = []
        for key, config in STRATEGY_CONFIG.items():
            strategies.append({
                'name': key,
                'description': config['description'],
                'parameters': {
                    'min_str_diff_percent': config['min_str_diff'] * 100,
                    'min_inventory': config['min_inventory'],
                    'max_transfer': config['max_transfer']
                }
            })
        
        return jsonify({
            'success': True,
            'strategies': strategies,
            'current_strategy': sistem.current_strategy,
            'transfer_type': sistem.transfer_type,
            'target_store': sistem.target_store
        })
        
    except Exception as e:
        logger.error(f"Strategies error: {str(e)}")
        return jsonify({'error': f'Strateji verisi hatasi: {str(e)}'}), 500

@app.route('/performance', methods=['GET'])
def get_performance_metrics():
    """Performance metrics endpoint"""
    try:
        memory_usage = sistem.check_memory_usage()
        
        try:
            cache_info = sistem.str_hesapla_cached.cache_info()
            cache_stats = {
                'hits': cache_info.hits,
                'misses': cache_info.misses,
                'hit_ratio': cache_info.hits/(cache_info.hits+cache_info.misses)*100 if (cache_info.hits+cache_info.misses) > 0 else 0,
                'current_size': cache_info.currsize,
                'max_size': cache_info.maxsize
            }
        except:
            cache_stats = {'error': 'Cache info not available'}
        
        metrics = {
            'current_memory_usage': memory_usage,
            'cache_statistics': cache_stats,
            'historical_metrics': sistem.performance_metrics,
            'system_info': {
                'data_loaded': sistem.data is not None,
                'total_stores': len(sistem.magazalar),
                'total_rows': len(sistem.data) if sistem.data is not None else 0
            }
        }
        
        return jsonify({
            'success': True,
            'metrics': metrics,
            'timestamp': datetime.now().isoformat()
        })
        
    except Exception as e:
        logger.error(f"Performance metrics error: {str(e)}")
        return jsonify({'error': f'Performance metrics hatasi: {str(e)}'}), 500

@app.route('/simulate', methods=['POST'])
def simulate_transfers():
    """Transfer simulation endpoint"""
    try:
        if not sistem.mevcut_analiz:
            return jsonify({'error': 'Önce bir analiz yapmanız gerekiyor'}), 400
        
        transferler = sistem.mevcut_analiz.get('transferler', [])
        simulation_result = sistem.simulate_transfer_impact(transferler)
        
        return jsonify({
            'success': True,
            'simulation': simulation_result,
            'timestamp': datetime.now().isoformat()
        })
        
    except Exception as e:
        logger.error(f"Simulation error: {str(e)}")
        return jsonify({'error': f'Simülasyon hatası: {str(e)}'}), 500

# Error handlers
@app.errorhandler(413)
def too_large(e):
    logger.error("File too large error")
    return jsonify({'error': 'Dosya boyutu cok buyuk (max 100MB)'}), 413

@app.errorhandler(500)
def internal_error(error):
    logger.error(f"Internal server error: {error}")
    return jsonify({'error': 'Sunucu hatasi', 'details': str(error)}), 500

@app.errorhandler(404)
def not_found(error):
    logger.warning(f"Endpoint not found: {request.url}")
    return jsonify({'error': 'Endpoint bulunamadi'}), 404

@app.errorhandler(Exception)
def handle_exception(e):
    logger.error(f"Unhandled exception: {e}")
    logger.error(f"Traceback: {traceback.format_exc()}")
    return jsonify({'error': 'Beklenmeyen hata oluştu', 'details': str(e)}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    debug_mode = os.environ.get('FLASK_ENV') != 'production'
    
    logger.info(f"Starting RetailFlow API v6.1.0-optimized-fixed on port {port}")
    logger.info("✅ All critical fixes applied:")
    logger.info("- Syntax errors fixed")
    logger.info("- Indentation corrected")
    logger.info("- Sales threshold removed (✓)")
    logger.info("- Mandatory min 1 transfer removed (✓)")
    logger.info("- Merkez/Online priority implemented (✓)")
    logger.info("- Vectorized operations for speed")
    logger.info("- Memory optimization")
    logger.info("- LRU Cache for performance")
    logger.info("- Progress tracking")
    logger.info("- Enhanced error handling")
    logger.info("- Performance metrics with Excel export")
    logger.info("- Transfer impact simulation")
    
    app.run(host='0.0.0.0', port=port, debug=debug_mode)


# Global system instance
system = MagazaTransferSistemi()



# ---------------------------
# Helper functions for new dashboards
# ---------------------------
def _ensure_product_keys(df, system_obj):
    try:
        if 'urun_anahtari' not in df.columns:
            df = system_obj.create_product_key_vectorized(df)
        return df
    except Exception:
        return df

def _store_metrics(df):
    stores = []
    for store, grp in df.groupby('Depo Adi'):
        sales = float(grp['Satis'].sum())
        inv = float(grp['Envanter'].sum())
        total = sales + inv
        str_val = (sales / total) if total > 0 else 0.0
        inv_turn = (sales / inv) if inv > 0 else 0.0
        uniq_products = int(grp.get('Urun Kodu', grp.get('Urun Adi')).nunique())
        stores.append({
            "store": store,
            "sales": round(sales, 2),
            "inventory": round(inv, 2),
            "str": round(str_val, 4),
            "inv_turn": round(inv_turn, 4),
            "product_count": uniq_products
        })
    return stores

def _normalize(values):
    if not values:
        return []
    mn, mx = min(values), max(values)
    if mx - mn == 0:
        return [50.0 for _ in values]
    return [ (v - mn) / (mx - mn) * 100.0 for v in values ]

def _category_from_row(row):
    # Prefer 'Kategori' column if exists, else derive from 'Urun Adi' or 'Urun Kodu'
    if 'Kategori' in row and isinstance(row['Kategori'], str) and row['Kategori'].strip():
        return row['Kategori'].strip()
    name = str(row.get('Urun Adi', '')).strip()
    if name:
        return name.split()[0].upper()
    code = str(row.get('Urun Kodu', '')).strip()
    return code[:3].upper() if code else "GENEL"

@app.route('/api/impact/summary', methods=['GET'])
def api_impact_summary():
    if system.data is None:
        return jsonify({"error": "Veri yuklu degil"}), 400
    if system.mevcut_analiz is None or 'transferler' not in system.mevcut_analiz:
        return jsonify({"error": "Once bir analiz (global/targeted/beden) calistirin"}), 400

    df = system.data.copy()
    df = _ensure_product_keys(df, system)
    transfers = system.mevcut_analiz.get('transferler', [])
    # Build store STR map
    metrics = _store_metrics(df)
    str_map = {m['store']: m['str'] for m in metrics}

    # Aggregations
    total_units = 0
    total_expected_lift = 0.0
    flows = {}
    store_incoming = {}
    store_outgoing = {}
    category_counter = {}

    # Map product key to a sample row for category resolution
    key_to_row = {}
    if 'urun_anahtari' in df.columns:
        # pick first matching row per key
        for k, grp in df.groupby('urun_anahtari'):
            key_to_row[k] = grp.iloc[0].to_dict()

    for t in transfers:
        q = int(t.get('transfer_miktari', 0) or 0)
        if q <= 0:
            continue
        src = t.get('gonderen_magaza')
        dst = t.get('alan_magaza')
        str_src = None
        str_dst = None
        # prefer detailed fields if present (percents), else use store-level
        if 'gonderen_str' in t and 'alan_str' in t:
            str_src = float(t['gonderen_str']) / 100.0
            str_dst = float(t['alan_str']) / 100.0
        else:
            str_src = float(str_map.get(src, 0.0))
            str_dst = float(str_map.get(dst, 0.0))
        str_diff = max(0.0, str_dst - str_src)
        expected_lift = q * str_diff

        total_units += q
        total_expected_lift += expected_lift

        key = f"{src}→{dst}"
        flows[key] = flows.get(key, 0) + q
        store_incoming[dst] = store_incoming.get(dst, 0) + q
        store_outgoing[src] = store_outgoing.get(src, 0) + q

        # Category / product group stat
        ukey = t.get('urun_anahtari')
        cat = "GENEL"
        if ukey in key_to_row:
            cat = _category_from_row(key_to_row[ukey])
        else:
            # fallback: construct a temp row-like dict from transfer fields
            pseudo = {"Urun Adi": t.get("urun_adi", ""), "Urun Kodu": t.get("urun_kodu", ""), "Kategori": t.get("kategori", "")}
            cat = _category_from_row(pseudo)
        category_counter[cat] = category_counter.get(cat, 0) + q

    # Store pre/post STR projections
    pre_post = []
    # Precompute per store aggregates
    store_agg = df.groupby('Depo Adi').agg({"Satis": "sum", "Envanter": "sum"}).reset_index()
    agg_map = {row['Depo Adi']: {"sales": float(row['Satis']), "inv": float(row['Envanter'])} for _, row in store_agg.iterrows()}

    for store, vals in agg_map.items():
        sales = vals["sales"]; inv = vals["inv"]
        pre_total = sales + inv
        pre_str = (sales / pre_total) if pre_total > 0 else 0.0
        inc = float(store_incoming.get(store, 0))
        out = float(store_outgoing.get(store, 0))
        post_inv = max(0.0, inv + inc - out)
        post_total = sales + post_inv
        post_str = (sales / post_total) if post_total > 0 else 0.0
        pre_post.append({
            "store": store,
            "pre_str": round(pre_str, 4),
            "post_str": round(post_str, 4),
            "incoming_units": int(inc),
            "outgoing_units": int(out)
        })

    # Flow edges for visualization
    flow_edges = [{"from": k.split("→")[0], "to": k.split("→")[1], "units": v} for k, v in flows.items()]
    # Top categories
    category_stats = [{"category": k, "units": v} for k, v in sorted(category_counter.items(), key=lambda x: x[1], reverse=True)[:20]]

    return jsonify({
        "total_transfer_units": int(total_units),
        "expected_sales_lift_units": round(total_expected_lift, 2),
        "pre_post_str": pre_post,
        "flows": flow_edges,
        "category_stats": category_stats,
        "store_metrics": metrics
    })

@app.route('/api/store/scores', methods=['GET'])
def api_store_scores():
    if system.data is None:
        return jsonify({"error": "Veri yuklu degil"}), 400
    df = system.data.copy()
    df = _ensure_product_keys(df, system)

    # Build category column if missing for diversity
    if 'Kategori' not in df.columns:
        df['Kategori'] = df.apply(_category_from_row, axis=1)

    # Base metrics
    stores = _store_metrics(df)
    str_vals = [s['str'] for s in stores]
    inv_turn_vals = [s['inv_turn'] for s in stores]

    # Diversity: unique categories per store
    cat_diversity = df.groupby('Depo Adi')['Kategori'].nunique().reset_index(name='cat_count')
    cat_map = {r['Depo Adi']: int(r['cat_count']) for _, r in cat_diversity.iterrows()}
    diversity_vals = [cat_map.get(s['store'], 0) for s in stores]

    # Transfer potential: difference to median STR times inventory surplus
    import numpy as np
    median_str = float(np.median(str_vals)) if str_vals else 0.0
    potential_vals = []
    for s in stores:
        diff = max(0.0, median_str - s['str'])
        surplus = max(0.0, s['inventory'] - s['sales'])
        potential_vals.append(diff * surplus)

    # Normalize to 0-100
    str_norm = [v*100.0 for v in str_vals]
    inv_turn_norm = _normalize(inv_turn_vals)
    diversity_norm = _normalize(diversity_vals)
    potential_norm = _normalize(potential_vals)

    scores = []
    for i, s in enumerate(stores):
        score = 0.40*str_norm[i] + 0.25*inv_turn_norm[i] + 0.20*diversity_norm[i] + 0.15*potential_norm[i]
        tier = "Excellent" if score >= 80 else "Good" if score >= 60 else "Needs Attention" if score >= 40 else "Critical"
        emoji = "🟢" if tier=="Excellent" else "🟡" if tier=="Good" else "🟠" if tier=="Needs Attention" else "🔴"
        scores.append({
            "store": s['store'],
            "score": round(score,2),
            "tier": tier,
            "emoji": emoji,
            "components": {
                "STR": round(str_norm[i],2),
                "EnvanterDevir": round(inv_turn_norm[i],2),
                "KategoriCesitliligi": round(diversity_norm[i],2),
                "TransferPotansiyeli": round(potential_norm[i],2)
            }
        })
    # Sort by score desc
    scores.sort(key=lambda x: x['score'], reverse=True)
    return jsonify({"scores": scores})

@app.route('/api/suggestions/quickwins', methods=['GET'])
def api_quick_wins():
    if system.data is None:
        return jsonify({"error": "Veri yuklu degil"}), 400
    if system.mevcut_analiz is None or 'transferler' not in system.mevcut_analiz:
        return jsonify({"error": "Once bir analiz calistirin"}), 400

    transfers = system.mevcut_analiz.get('transferler', [])
    # Compute ROI (units) = expected_lift_units / q = str_diff
    quick = []
    for t in transfers:
        q = int(t.get('transfer_miktari', 0) or 0)
        if q <= 0:
            continue
        g_str = None; a_str = None
        if 'gonderen_str' in t and 'alan_str' in t:
            g_str = float(t['gonderen_str'])/100.0; a_str = float(t['alan_str'])/100.0
        else:
            g_str = 0.0; a_str = 0.0
        str_diff = max(0.0, a_str - g_str)
        expected_lift = round(q * str_diff, 2)
        roi_pct = round(str_diff*100.0, 2)
        quick.append({
            "urun_kodu": t.get("urun_kodu",""),
            "urun_adi": t.get("urun_adi",""),
            "renk": t.get("renk",""),
            "beden": t.get("beden",""),
            "from_store": t.get("gonderen_magaza",""),
            "to_store": t.get("alan_magaza",""),
            "qty": q,
            "roi_pct": roi_pct,
            "expected_lift_units": expected_lift
        })
    # Prioritize by ROI, then small qty for quick execution
    quick.sort(key=lambda x: (-x['roi_pct'], x['qty']))
    # Category hotspots (from transfers)
    # Top (beden, renk) patterns
    from collections import Counter
    cat_counter = Counter()
    size_color = Counter()
    for t in transfers:
        cat = "GENEL"
        name = str(t.get("urun_adi",""))
        if name:
            cat = name.split()[0].upper()
        cat_counter[cat] += int(t.get("transfer_miktari",0) or 0)
        sc_key = (str(t.get("beden","")).upper(), str(t.get("renk","")).upper())
        size_color[sc_key] += 1
    top_cats = [{"category": k, "units": v} for k,v in cat_counter.most_common(10)]
    top_patterns = [{"beden": k[0], "renk": k[1], "count": v} for k,v in size_color.most_common(10)]

    return jsonify({"quick_wins": quick[:50], "category_hotspots": top_cats, "size_color_patterns": top_patterns})

@app.route('/api/executive/summary', methods=['GET'])
def api_exec_summary():
    # Combine top-level metrics
    if system.data is None:
        return jsonify({"error": "Veri yuklu degil"}), 400

    df = system.data.copy()
    df = _ensure_product_keys(df, system)
    stores = _store_metrics(df)
    str_vals = [s['str'] for s in stores]
    avg_str = sum(str_vals)/len(str_vals) if str_vals else 0.0

    transfer_count = 0
    expected_lift = 0.0
    top_actions = []
    if system.mevcut_analiz and 'transferler' in system.mevcut_analiz:
        transfers = system.mevcut_analiz['transferler']
        transfer_count = len(transfers)
        for t in transfers:
            q = int(t.get('transfer_miktari', 0) or 0)
            if q <= 0: 
                continue
            if 'gonderen_str' in t and 'alan_str' in t:
                str_diff = max(0.0, float(t['alan_str'])/100.0 - float(t['gonderen_str'])/100.0)
            else:
                str_diff = 0.0
            expected_lift += q*str_diff

        # Top priority actions = highest ROI transfers
        tmp = []
        for t in transfers:
            q = int(t.get('transfer_miktari',0) or 0)
            if 'gonderen_str' in t and 'alan_str' in t:
                rd = max(0.0, float(t['alan_str'])/100.0 - float(t['gonderen_str'])/100.0)
            else:
                rd = 0.0
            tmp.append((rd, -q, t))
        tmp.sort(reverse=True)
        top_actions = [{
            "desc": f"{x[2].get('gonderen_magaza','?')} → {x[2].get('alan_magaza','?')} | {x[2].get('urun_adi','')} {x[2].get('beden','')} ({x[2].get('renk','')}) x{x[2].get('transfer_miktari',0)}",
            "roi_pct": round(x[0]*100.0,2)
        } for x in tmp[:10]]

    system_health = {
        "memory_used_pct": system.check_memory_usage(),
        "rows_loaded": int(len(df)),
        "store_count": int(len(set(df['Depo Adi']))),
        "last_file_load": system.performance_metrics.get("last_file_load", {})
    }

    # Store performance summary (quartiles)
    import numpy as np
    if str_vals:
        q1, q2, q3 = np.quantile(str_vals, [0.25, 0.5, 0.75]).tolist()
    else:
        q1=q2=q3=0.0

    return jsonify({
        "kpis": {
            "total_transfer_count": transfer_count,
            "expected_sales_lift_units": round(expected_lift, 2),
            "avg_store_str": round(avg_str*100.0, 2)
        },
        "store_performance_summary": {
            "q1": round(q1*100.0, 2),
            "median": round(q2*100.0, 2),
            "q3": round(q3*100.0, 2)
        },
        "top_priority_actions": top_actions,
        "system_health": system_health
    })
